from openpyxl import load_workbook
from Database_SQL.Repositories.Class_Repository import ClassRepository
from Database_SQL.Repositories.Student_Repository import StudentRepository
from Database_SQL.Repositories.Categories_Repository import CategoriesRepository
from Database_SQL.Repositories.Subjects_Repository import SubjectRepository
from Database_SQL.Repositories.Grades_Repository import GradesRepository
from Database_SQL.Repositories.Repository_Functions import RepositoryFunctions
from Database_SQL.Repositories.Junction_Table import JunctionTable
from Database_SQL.Features.DatabaseManager import DatabaseManager
from GUI_pyqt6.Configuration.Configuration import load_configuration
import os


def prepare_excel_sheet(excel_file_path, workbook, sheet):
    cell1, cell2 = sheet['A2'], sheet['B2']
    cell1.value, cell2.value = 'ID', 'Names'
    workbook.save(excel_file_path)


def extract_sheet_name(workbook):
    sheet_name = {}
    data = workbook.sheetnames[0].split("-")
    name = data[0].split(" ")
    sheet_name["Quarter"] = name[0].strip()
    sheet_name["Category"] = name[1].strip()
    sheet_name["Subject"] = data[1].strip()
    sheet_name["Class"] = data[2].strip()
    return sheet_name


def add_students(st_repo, sheet, class_name):
    for cell1, cell2 in zip(sheet['A'][2:], sheet['B'][2:]):
        st_repo.create(cell1.value, cell2.value, class_name)


def add_categories(j_table, sheet, quarter, subject_name):
    for column in sheet.iter_cols(min_row=2, max_row=2, min_col=3):
        for cell in column:
            category_name = cell.value.replace(quarter, "").strip()
            j_table.create_category_subject(category_name.lower(), subject_name)


def add_grades(g_repo, sheet, subject_name, quarter, class_name, r_func):
    class_id = r_func.get_class_by_class_name(class_name).id
    index = 2
    for row in sheet.iter_rows(min_row=3):
        student_id = row[0].value

        for column in sheet.iter_cols(min_col=3):
            category_name = column[1].value.replace(quarter, "").strip()
            g_repo.create(student_id, class_id, category_name, subject_name, quarter[1], column[index].value)

        index += 1


def add_gradebook_to_database(excel_file_path, database_path):
    dm = DatabaseManager(database_path)

    r_func = RepositoryFunctions(dm)
    s_repo = SubjectRepository(dm, r_func)
    st_repo = StudentRepository(dm)
    c_repo = ClassRepository(dm, r_func)
    cat_repo = CategoriesRepository(dm)
    g_repo = GradesRepository(dm, r_func)
    j_table = JunctionTable(dm, c_repo, s_repo, cat_repo, r_func)
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    prepare_excel_sheet(excel_file_path, workbook, sheet)
    sheet_name = extract_sheet_name(workbook)

    j_table.create_class_subject(sheet_name["Class"], sheet_name["Subject"])

    add_students(st_repo, sheet, sheet_name["Class"])

    add_categories(j_table, sheet, sheet_name["Quarter"], sheet_name["Subject"])

    add_grades(g_repo, sheet, sheet_name["Subject"], sheet_name["Quarter"], sheet_name["Class"], r_func)


def import_excel_files_from_directory(directory_path):
    pathnames = []
    with os.scandir(directory_path) as entries:
        for entry in entries:
            if entry.is_file() \
                    and not entry.name.startswith('~$') \
                    and not entry.name.startswith('.') \
                    and entry.name.endswith('.xlsx') \
                    or entry.name.endswith('.xls'):
                pathnames.append(entry.path)

    settings = load_configuration("Configuration/settings.json")
    database_path = settings.database_path

    for pathname in pathnames:
        add_gradebook_to_database(pathname, database_path)
